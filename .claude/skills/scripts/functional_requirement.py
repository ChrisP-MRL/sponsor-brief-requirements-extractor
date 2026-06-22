#!/usr/bin/env python3
"""
Functional Requirements Extractor

This script reads text files, extracts functional requirements using AI/NLP,
and populates them into a copy of the Functional Requirements Specification Excel template.
"""

import os
import sys
import re
import shutil
from pathlib import Path
from datetime import datetime
from typing import List, Dict, Tuple
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill


class FunctionalRequirementExtractor:
    def __init__(self, template_path: str, output_dir: str = "requirements"):
        self.template_path = Path(template_path)
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)

        if not self.template_path.exists():
            raise FileNotFoundError(f"Template not found: {self.template_path}")

    def extract_requirements_from_text(self, text: str) -> List[Dict[str, str]]:
        """
        Extract functional requirements from text content.
        Looks for patterns that indicate requirements:
        - Numbered lists (1., 2., etc.)
        - "Must", "Should", "Shall" statements
        - User story format ("As a... I want... So that...")
        - Feature descriptions
        """
        requirements = []

        # Pattern 1: Explicit requirement keywords
        req_patterns = [
            r'(?:must|shall|should|will)\s+(?:be able to|allow|provide|support|enable|implement|include).*?(?:\.|$)',
            r'(?:the system|user|application|software)\s+(?:must|shall|should|will).*?(?:\.|$)',
            r'(?:requirement|req|fr)[\s\-:]+\d+.*?(?:\n|$)',
        ]

        for pattern in req_patterns:
            matches = re.finditer(pattern, text, re.IGNORECASE | re.MULTILINE)
            for match in matches:
                req_text = match.group(0).strip()
                if len(req_text) > 20:  # Filter out very short matches
                    requirements.append({
                        'description': req_text,
                        'priority': self._determine_priority(req_text),
                        'category': self._determine_category(req_text)
                    })

        # Pattern 2: User stories format
        user_story_pattern = r'(?:as\s+a|as\s+an)\s+.*?(?:i\s+want|i\s+need).*?(?:so\s+that)?.*?(?:\n|$)'
        matches = re.finditer(user_story_pattern, text, re.IGNORECASE | re.MULTILINE)
        for match in matches:
            req_text = match.group(0).strip()
            if len(req_text) > 20:
                requirements.append({
                    'description': req_text,
                    'priority': 'Medium',
                    'category': 'User Story'
                })

        # Pattern 3: Numbered or bulleted lists that look like requirements
        lines = text.split('\n')
        for i, line in enumerate(lines):
            line = line.strip()
            # Check if line starts with number, bullet, or dash
            if re.match(r'^(?:\d+\.|\*|\-|\•)\s+', line):
                # Check if it contains requirement-like language
                if any(keyword in line.lower() for keyword in ['must', 'should', 'shall', 'will', 'enable', 'allow', 'provide', 'support']):
                    req_text = re.sub(r'^(?:\d+\.|\*|\-|\•)\s+', '', line)
                    if len(req_text) > 20:
                        requirements.append({
                            'description': req_text,
                            'priority': self._determine_priority(req_text),
                            'category': self._determine_category(req_text)
                        })

        # Remove duplicates based on similar text
        unique_requirements = []
        seen_texts = set()
        for req in requirements:
            # Normalize text for comparison
            normalized = req['description'].lower()[:100]
            if normalized not in seen_texts:
                seen_texts.add(normalized)
                unique_requirements.append(req)

        return unique_requirements

    def _determine_priority(self, text: str) -> str:
        """Determine priority based on keywords in the requirement text."""
        text_lower = text.lower()
        if 'must' in text_lower or 'critical' in text_lower or 'essential' in text_lower:
            return 'High'
        elif 'should' in text_lower or 'important' in text_lower:
            return 'Medium'
        elif 'could' in text_lower or 'optional' in text_lower or 'nice to have' in text_lower:
            return 'Low'
        return 'Medium'

    def _determine_category(self, text: str) -> str:
        """Determine requirement category based on content."""
        text_lower = text.lower()

        categories = {
            'Authentication': ['login', 'authentication', 'password', 'user account', 'sign in', 'sign up'],
            'Data Management': ['data', 'database', 'storage', 'save', 'retrieve', 'update', 'delete'],
            'User Interface': ['ui', 'interface', 'display', 'view', 'screen', 'button', 'form'],
            'Integration': ['integrate', 'api', 'external', 'third-party', 'interface with'],
            'Reporting': ['report', 'dashboard', 'analytics', 'metrics', 'export'],
            'Security': ['security', 'secure', 'encryption', 'authorization', 'permission', 'access control'],
            'Performance': ['performance', 'speed', 'fast', 'response time', 'scalable'],
            'Notification': ['notification', 'alert', 'email', 'notify', 'message'],
        }

        for category, keywords in categories.items():
            if any(keyword in text_lower for keyword in keywords):
                return category

        return 'General'

    def _split_category(self, category: str) -> Tuple[str, str]:
        """
        Split category into main category and sub-category.
        Maps technical categories to business-aligned main categories.
        """
        # Category mapping: technical category -> (main category, sub-category)
        category_mapping = {
            'Authentication': ('Security', 'Authentication'),
            'Data Management': ('Data', 'Data Management'),
            'User Interface': ('User Interface', 'Interface'),
            'Integration': ('Integration', 'System Integration'),
            'Reporting': ('Reporting', 'Analytics'),
            'Security': ('Security', 'Security Controls'),
            'Performance': ('Performance', 'System Performance'),
            'Notification': ('Communication', 'Notifications'),
            'General': ('General', 'General Requirements'),
            'User Story': ('Functional', 'User Stories'),
        }

        if category in category_mapping:
            return category_mapping[category]

        # Default fallback
        return ('General', category)

    def _sort_requirements_by_category(self, requirements: List[Dict[str, str]]) -> List[Dict[str, str]]:
        """
        Sort requirements by CATEGORY (main) and then SUB-CATEGORY.
        This ensures IDs are assigned in a logical order.
        """
        # Add category and sub-category to each requirement for sorting
        requirements_with_categories = []
        for req in requirements:
            main_category, sub_category = self._split_category(req['category'])
            requirements_with_categories.append({
                'req': req,
                'main_category': main_category,
                'sub_category': sub_category
            })

        # Sort by main category first, then by sub-category
        sorted_reqs = sorted(
            requirements_with_categories,
            key=lambda x: (x['main_category'], x['sub_category'])
        )

        # Return just the requirement objects
        return [item['req'] for item in sorted_reqs]

    def read_text_file(self, file_path: str) -> str:
        """Read content from a text file."""
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                return f.read()
        except UnicodeDecodeError:
            # Try with different encoding if UTF-8 fails
            with open(file_path, 'r', encoding='latin-1') as f:
                return f.read()

    def populate_excel(self, requirements: List[Dict[str, str]],
                      project_name: str = "Project",
                      source_file: str = "") -> str:
        """
        Populate the Excel template with extracted requirements.
        Returns the path to the created file.
        """
        # Copy template to output directory
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_filename = f"FRS_{project_name}_{timestamp}.xlsx"
        output_path = self.output_dir / output_filename

        shutil.copy(self.template_path, output_path)

        # Load the workbook
        wb = load_workbook(output_path)

        # Clean up problematic external references and drawings from template
        # These cause Excel corruption warnings when template has broken links
        self._clean_external_references(wb)

        # Assuming the first sheet contains the requirements table
        # You may need to adjust sheet name based on actual template
        if 'Requirements' in wb.sheetnames:
            ws = wb['Requirements']
        elif 'Functional Requirements' in wb.sheetnames:
            ws = wb['Functional Requirements']
        else:
            ws = wb.active

        # Find the starting row for requirements (skip headers)
        # Look for header row containing "Requirement" or "ID"
        start_row = self._find_header_row(ws) + 1

        # Sort requirements by category and sub-category before assigning IDs
        sorted_requirements = self._sort_requirements_by_category(requirements)

        # Populate requirements - simply write data to cells
        # Template may have some pre-formatted rows, but we'll just overwrite/extend
        for idx, req in enumerate(sorted_requirements, start=1):
            row = start_row + idx - 1

            # Template structure based on PRJ-Functional Requirements Specification:
            # Column A: CATEGORY (main functional area)
            # Column B: SUB-CATEGORY (detailed category)
            # Column C: ID (requirement identifier)
            # Column D: DESCRIPTION (requirement text)
            # Column E: MRL Priority (priority level)
            # Column F: COMPLIANCE (left blank - to be filled manually)

            # Determine main category and sub-category
            main_category, sub_category = self._split_category(req['category'])

            # Populate data
            ws.cell(row=row, column=1, value=main_category)  # CATEGORY
            ws.cell(row=row, column=2, value=sub_category)  # SUB-CATEGORY
            ws.cell(row=row, column=3, value=f"FR-{idx:03d}")  # ID (assigned after sorting)
            ws.cell(row=row, column=4, value=req['description'])  # DESCRIPTION
            ws.cell(row=row, column=5, value=req['priority'])  # MRL Priority
            # Column F (COMPLIANCE) is intentionally left blank

            # Apply basic text wrapping for readability
            ws.cell(row=row, column=4).alignment = Alignment(wrap_text=True, vertical='top')

        # Update metadata if template has a metadata sheet or section
        self._update_metadata(wb, project_name, source_file, len(requirements))

        # Save the workbook
        wb.save(output_path)
        return str(output_path)

    def _clean_external_references(self, wb):
        """
        Remove external formula references, named ranges, and drawings that cause corruption.
        The template may contain broken external links that Excel flags on open.
        """
        # Remove external links (these cause "External formula reference" errors)
        if hasattr(wb, '_external_links') and wb._external_links:
            wb._external_links.clear()

        # Remove named ranges that may reference external sources
        if hasattr(wb, 'defined_names'):
            # Get list of all defined name keys
            names_to_remove = list(wb.defined_names.definedName.keys()) if hasattr(wb.defined_names, 'definedName') else []

            # If the above doesn't work, try iterating directly
            if not names_to_remove:
                try:
                    names_to_remove = [name for name in wb.defined_names]
                except:
                    pass

            # Remove each named range
            for name in names_to_remove:
                try:
                    del wb.defined_names[name]
                except:
                    pass

        # Clean up drawings and charts from all worksheets
        for ws in wb.worksheets:
            # Remove drawings
            if hasattr(ws, '_drawing'):
                ws._drawing = None
            if hasattr(ws, '_drawings'):
                ws._drawings = []

            # Remove charts
            if hasattr(ws, '_charts'):
                ws._charts = []

    def _find_header_row(self, ws) -> int:
        """Find the row containing headers like 'Requirement', 'ID', etc."""
        for row_idx in range(1, 20):  # Check first 20 rows
            for cell in ws[row_idx]:
                if cell.value and isinstance(cell.value, str):
                    if any(header in cell.value.lower() for header in ['requirement', 'id', 'description']):
                        return row_idx
        return 1  # Default to row 1 if no header found

    def _update_metadata(self, wb, project_name: str, source_file: str, req_count: int):
        """Update metadata in the workbook if metadata section exists."""
        # Try to find and update metadata sheet or cells
        if 'Metadata' in wb.sheetnames:
            ws = wb['Metadata']
            # Update common metadata fields
            metadata_fields = {
                'Project Name': project_name,
                'Source File': source_file,
                'Generated Date': datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                'Total Requirements': req_count
            }

            for row_idx in range(1, 30):
                cell_a = ws.cell(row=row_idx, column=1)
                if cell_a.value and isinstance(cell_a.value, str):
                    for field_name, field_value in metadata_fields.items():
                        if field_name.lower() in cell_a.value.lower():
                            ws.cell(row=row_idx, column=2, value=field_value)

    def process_file(self, input_file: str, project_name: str = None) -> str:
        """
        Main method to process a text file and generate requirements Excel.

        Args:
            input_file: Path to the text file containing requirements
            project_name: Name of the project (defaults to filename)

        Returns:
            Path to the generated Excel file
        """
        input_path = Path(input_file)

        if not input_path.exists():
            raise FileNotFoundError(f"Input file not found: {input_file}")

        # Determine project name
        if project_name is None:
            project_name = input_path.stem

        # Read and extract requirements
        print(f"Reading file: {input_file}")
        text_content = self.read_text_file(input_file)

        print("Extracting functional requirements...")
        requirements = self.extract_requirements_from_text(text_content)

        print(f"Found {len(requirements)} functional requirements")

        if len(requirements) == 0:
            print("Warning: No requirements found. The file may not contain requirement-like statements.")
            print("Tip: Requirements should contain keywords like 'must', 'should', 'shall', 'will', etc.")

        # Populate Excel
        print("Generating Excel file...")
        output_path = self.populate_excel(requirements, project_name, str(input_path))

        print(f"Successfully created: {output_path}")
        return output_path


def main():
    """Command-line interface for the script."""
    if len(sys.argv) < 2:
        print("Usage: python functional_requirement.py <input_text_file> [project_name]")
        print("\nExample:")
        print("  python functional_requirement.py requirements.txt MyProject")
        sys.exit(1)

    input_file = sys.argv[1]
    project_name = sys.argv[2] if len(sys.argv) > 2 else None

    # Determine template path relative to script location
    script_dir = Path(__file__).parent
    template_path = script_dir.parent / "code-review" / "templates" / "PRJ-Functional Requirements Specification.xlsx"

    if not template_path.exists():
        # Try alternative path
        template_path = Path(".claude/skills/code-review/templates/PRJ-Functional Requirements Specification.xlsx")

    try:
        extractor = FunctionalRequirementExtractor(str(template_path))
        output_file = extractor.process_file(input_file, project_name)
        print(f"\nRequirements extracted successfully!")
        print(f"Output: {output_file}")
    except Exception as e:
        print(f"Error: {e}", file=sys.stderr)
        sys.exit(1)


if __name__ == "__main__":
    main()
