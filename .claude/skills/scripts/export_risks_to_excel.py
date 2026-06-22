#!/usr/bin/env python3
"""
Export risks from markdown format to Excel with proper formatting.
"""

import sys
import re
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter


def parse_risk_table(markdown_content):
    """Parse risk table from markdown content."""
    risks = []

    # Find the Risks section
    risk_section_match = re.search(r'## Risks\s*\n\s*\|.*?\|.*?\n\|[-\s|]+\n((?:\|.*?\n)+)',
                                   markdown_content, re.MULTILINE)

    if not risk_section_match:
        print("Warning: No Risks section found in markdown")
        return risks

    table_content = risk_section_match.group(1)

    # Parse each row
    for line in table_content.strip().split('\n'):
        if not line.strip() or line.strip() == '|':
            continue

        # Split by | and clean up
        columns = [col.strip() for col in line.split('|')]
        # Remove empty first/last elements from splitting
        columns = [col for col in columns if col]

        if len(columns) >= 4:
            risk_text = columns[0]
            impact = columns[1]
            likelihood = columns[2]
            mitigation = columns[3]

            # Skip placeholder rows
            if '[TO BE COMPLETED]' in risk_text or not risk_text:
                continue

            # Determine risk category based on keywords
            category = categorize_risk(risk_text)

            risks.append({
                'risk': risk_text,
                'impact': impact.upper() if impact else 'M',
                'likelihood': likelihood.upper() if likelihood else 'M',
                'mitigation': mitigation,
                'category': category
            })

    return risks


def categorize_risk(risk_text):
    """Categorize risk based on keywords in the risk description."""
    risk_lower = risk_text.lower()

    if any(word in risk_lower for word in ['technical', 'integration', 'system', 'technology', 'feasibility']):
        return 'Technical'
    elif any(word in risk_lower for word in ['resource', 'staff', 'personnel', 'capability', 'capacity']):
        return 'Resource'
    elif any(word in risk_lower for word in ['schedule', 'timeline', 'delay', 'deadline', 'dependency']):
        return 'Schedule'
    elif any(word in risk_lower for word in ['budget', 'cost', 'financial', 'funding']):
        return 'Budget'
    elif any(word in risk_lower for word in ['stakeholder', 'communication', 'buy-in', 'resistance', 'change']):
        return 'Stakeholder'
    elif any(word in risk_lower for word in ['regulatory', 'compliance', 'external', 'supplier', 'market']):
        return 'External'
    else:
        return 'Other'


def sort_risks(risks):
    """Sort risks by Impact (H→M→L) then Likelihood (H→M→L)."""
    priority_map = {'H': 1, 'M': 2, 'L': 3}

    return sorted(risks, key=lambda r: (
        priority_map.get(r['impact'], 2),
        priority_map.get(r['likelihood'], 2)
    ))


def export_to_excel(risks, output_path, project_name):
    """Export risks to Excel with formatting."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Risks"

    # Define styles
    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    cell_font = Font(size=10)
    cell_alignment = Alignment(vertical="top", wrap_text=True)

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Add title
    ws.merge_cells('A1:E1')
    title_cell = ws['A1']
    title_cell.value = f"Risk Register: {project_name}"
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Add headers
    headers = ['Risk', 'Impact', 'Likelihood', 'Mitigation', 'Risk Category']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Add data
    for row_num, risk in enumerate(risks, 3):
        ws.cell(row=row_num, column=1, value=risk['risk'])
        ws.cell(row=row_num, column=2, value=risk['impact'])
        ws.cell(row=row_num, column=3, value=risk['likelihood'])
        ws.cell(row=row_num, column=4, value=risk['mitigation'])
        ws.cell(row=row_num, column=5, value=risk['category'])

        # Apply formatting to all cells in row
        for col_num in range(1, 6):
            cell = ws.cell(row=row_num, column=col_num)
            cell.font = cell_font
            cell.alignment = cell_alignment
            cell.border = thin_border

    # Auto-size columns
    column_widths = {
        'A': 40,  # Risk
        'B': 10,  # Impact
        'C': 12,  # Likelihood
        'D': 45,  # Mitigation
        'E': 15   # Risk Category
    }

    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # Set row heights
    ws.row_dimensions[1].height = 25
    ws.row_dimensions[2].height = 30

    # Save workbook
    wb.save(output_path)
    print(f"[OK] Risks exported to: {output_path}")
    print(f"  Total risks: {len(risks)}")


def main():
    if len(sys.argv) < 3:
        print("Usage: python export_risks_to_excel.py <markdown_file> <project_name>")
        sys.exit(1)

    markdown_file = Path(sys.argv[1])
    project_name = sys.argv[2]

    if not markdown_file.exists():
        print(f"Error: Markdown file not found: {markdown_file}")
        sys.exit(1)

    # Read markdown content
    with open(markdown_file, 'r', encoding='utf-8') as f:
        content = f.read()

    # Parse risks
    risks = parse_risk_table(content)

    if not risks:
        print("Warning: No risks found to export")
        sys.exit(0)

    # Sort risks
    risks = sort_risks(risks)

    # Create output directory if needed
    output_dir = Path("Output")
    output_dir.mkdir(exist_ok=True)

    # Export to Excel
    output_file = output_dir / f"Risks-{project_name}.xlsx"
    export_to_excel(risks, output_file, project_name)


if __name__ == "__main__":
    main()
